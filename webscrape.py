# Standard library imports
import os
import sys
import re
import logging
from datetime import datetime

# Third-party imports
import requests      # For HTTP requests to download CSV files
import pandas as pd  # For data manipulation and Excel export
from openpyxl.styles import Alignment  # For Excel formatting
# tzlocal is used later for timezone handling (see below)

def setup_directories(script_dir):
    """
    Ensure the input and output directories exist within the given script directory.

    Args:
        script_dir (str): The absolute path to the directory containing the script or executable.

    Returns:
        tuple: (input_folder, output_folder) - Absolute paths to the input and output directories.

    Notes:
        - Creates the directories if they do not already exist.
        - Used to organize input CSVs and output Excel files for the webscrape process.
    """
    input_folder = os.path.join(script_dir, "input")
    output_folder = os.path.join(script_dir, "output")
    os.makedirs(input_folder, exist_ok=True)
    os.makedirs(output_folder, exist_ok=True)
    return input_folder, output_folder

def read_url_from_file(script_dir):
    """
    Reads the download URL from the configuration file 'sam_url.txt' in the cfg directory.

    Args:
        script_dir (str): The absolute path to the directory containing the script or executable.

    Returns:
        str or None: The URL string if successfully read, otherwise None.

    Notes:
        - Expects the file 'cfg/sam_url.txt' to exist and contain the URL on the first line.
        - Logs an info message and returns None if the file cannot be read.
    """
    cfg_dir = os.path.join(script_dir, "cfg")
    url_file = os.path.join(cfg_dir, "sam_url.txt")
    try:
        with open(url_file, "r") as uf:
            url = uf.readline().strip()
        return url
    except Exception as e:
        logging.info(f"Failed to read URL from {url_file}: {e}")
        return None

def download_csv_file(url, csv_filename):
    """
    Downloads a CSV file from the specified URL and saves it to the given filename.

    Args:
        url (str): The URL to download the CSV file from.
        csv_filename (str): The local file path where the downloaded CSV will be saved.

    Returns:
        None

    Notes:
        - Logs progress and errors using the logging module.
        - Uses streaming download to handle large files efficiently.
        - Overwrites the file if it already exists.
    """
    try:
        logging.info(f"Downloading CSV from {url} ...")
        with requests.get(url, stream=True, timeout=60) as response:
            response.raise_for_status()
            with open(csv_filename, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        logging.info(f"File downloaded and saved to: {csv_filename}")
    except Exception as e:
        logging.info(f"Failed to download file: {e}")

def make_pattern_and_flags(k):
    """
    Generate a regex pattern and flags for matching a keyword or organization string.

    Args:
        k (str): The keyword or organization string to match.

    Returns:
        tuple: (pattern, flags) where pattern is a regex string and flags is an int for re module flags.

    Notes:
        - If the string contains spaces, it matches as a whole word (case-insensitive).
        - If the string is all uppercase letters, it matches as a whole word (case-sensitive).
        - Otherwise, matches as a whole word (case-insensitive).
    """
    if " " in k:
        pattern = r'(?<!\w)' + re.escape(k) + r'(?!\w)'
        flags = re.IGNORECASE
    elif k.isalpha() and k.isupper():
        pattern = r'\b' + re.escape(k) + r'\b'
        flags = 0
    else:
        pattern = r'\b' + re.escape(k) + r'\b'
        flags = re.IGNORECASE
    return pattern, flags

def filter_by_organization(df, keyorgs):
    """
    Filter DataFrame rows where the 'Sub-Tier' column matches any organization in the provided list.

    Args:
        df (pd.DataFrame): The input DataFrame containing a 'Sub-Tier' column.
        keyorgs (list of str): List of organization names to match against.

    Returns:
        pd.DataFrame: Filtered DataFrame containing only rows where 'Sub-Tier' matches any organization.

    Notes:
        - Uses regex patterns and flags from make_pattern_and_flags for robust matching.
        - If keyorgs is empty, returns the original DataFrame unchanged.
    """
    if not keyorgs:
        return df
    org_masks = []
    for org in keyorgs:
        pattern, flags = make_pattern_and_flags(org)
        mask = df['Sub-Tier'].str.contains(pattern, case=False, na=False, regex=True, flags=flags)
        org_masks.append(mask)
    if org_masks:
        combined_org_mask = org_masks[0]
        for m in org_masks[1:]:
            combined_org_mask = combined_org_mask | m
        return df[combined_org_mask].copy()
    return df

def filter_by_keywords(df, keywords):
    """
    Filter DataFrame rows where the 'Description' column matches any keyword in the provided list.

    Args:
        df (pd.DataFrame): The input DataFrame containing a 'Description' column.
        keywords (list of str): List of keywords to match against.

    Returns:
        pd.DataFrame: Filtered DataFrame containing only rows where 'Description' matches any keyword.
        Adds 'Matched Keywords' and 'Matched Keywords Count' columns to the result.

    Notes:
        - Uses regex patterns and flags from make_pattern_and_flags for robust keyword matching.
        - If keywords is empty, returns the original DataFrame with an empty 'Matched Keywords' column.
        - The result is sorted by number of matched keywords (descending) and then alphabetically.
    """
    if not keywords:
        df = df.copy()
        df.insert(0, "Matched Keywords", [""] * len(df))
        return df
    org_indices = list(df.index)
    matched_keywords_per_row = {idx: [] for idx in org_indices}
    for k in keywords:
        pattern, flags = make_pattern_and_flags(k)
        mask = df['Description'].str.contains(pattern, case=False, na=False, regex=True, flags=flags)
        for idx, is_match in zip(org_indices, mask):
            if is_match:
                matched_keywords_per_row[idx].append(k)
    matched_mask = df.index.map(lambda idx: len(matched_keywords_per_row[idx]) > 0)
    filtered_data = df[matched_mask].copy()
    filtered_indices = filtered_data.index.tolist()
    matched_keywords_filtered = [", ".join(matched_keywords_per_row[i]) for i in filtered_indices]
    matched_keywords_count = [len(matched_keywords_per_row[i]) for i in filtered_indices]
    filtered_data.insert(0, "Matched Keywords", matched_keywords_filtered)
    filtered_data.insert(1, "Matched Keywords Count", matched_keywords_count)
    filtered_data = filtered_data.sort_values(
        by=["Matched Keywords Count", "Matched Keywords"],
        ascending=[False, True]
    ).reset_index(drop=True)
    return filtered_data

def filter_columns(df, keep_cols):
    """
    Filter DataFrame columns to retain only those specified in keep_cols, always keeping 'Matched Keywords' and 'Matched Keywords Count' first.

    Args:
        df (pd.DataFrame): The input DataFrame to filter columns from.
        keep_cols (list of str): List of column names to retain in the output.

    Returns:
        pd.DataFrame: DataFrame with only the specified columns, with 'Matched Keywords' and 'Matched Keywords Count' as the first columns if present.

    Notes:
        - If 'Matched Keywords' or 'Matched Keywords Count' are not in keep_cols, they are prepended to the output columns if present in df.
        - Columns not present in df are ignored.
    """
    cols_to_keep = [col for col in keep_cols if col in df.columns]
    for special_col in ["Matched Keywords", "Matched Keywords Count"]:
        if special_col not in cols_to_keep:
            cols_to_keep = [special_col] + cols_to_keep
        else:
            cols_to_keep = [special_col] + [col for col in cols_to_keep if col != special_col]
    return df[cols_to_keep]

def clean_dataframe(df):
    """
    Sanitize DataFrame column names and remove non-printable/control characters from all string columns.

    Args:
        df (pd.DataFrame): The input DataFrame to clean.

    Returns:
        pd.DataFrame: Cleaned DataFrame with sanitized column names and cleaned string values.

    Notes:
        - Column names are truncated to 31 characters and illegal characters are replaced with underscores.
        - All string columns have control/non-printable characters removed.
    """
    illegal_chars = [':', '\\', '/', '?', '*', '[', ']']
    def sanitize(val):
        val = str(val)
        for ch in illegal_chars:
            val = val.replace(ch, '_')
        return val[:31]
    df.columns = [sanitize(col) for col in df.columns]
    def clean_string(s):
        if isinstance(s, str):
            return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', s)
        return s
    for col in df.select_dtypes(include=['object']).columns:
        df.loc[:, col] = df[col].apply(clean_string)
    return df

def write_to_excel(filtered_data, keyword_counts_df, output_filename):
    """
    Write filtered data and keyword counts to an Excel file, format columns, and add hyperlinks.

    Args:
        filtered_data (pd.DataFrame): DataFrame of filtered keyword matches to write to Excel.
        keyword_counts_df (pd.DataFrame): DataFrame of keyword match counts to write to Excel.
        output_filename (str): Path to the output Excel file.

    Returns:
        None

    Notes:
        - Writes two sheets: 'Keyword Matches' and 'Keyword Match Counts'.
        - Formats columns for readability and auto-width.
        - Adds hyperlinks to cells in the 'Link' column if values are valid URLs.
    """
    from openpyxl.styles import Alignment
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        filtered_data.to_excel(writer, index=False, sheet_name='Keyword Matches')
        keyword_counts_df.to_excel(writer, index=False, sheet_name='Keyword Match Counts')
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
                cell.alignment = Alignment(horizontal='left')
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col[1:]:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                header_length = len(str(col[0].value)) if col[0].value is not None else 0
                best_length = max(max_length, header_length) + 2
                worksheet.column_dimensions[col_letter].width = best_length
            for col_idx, cell in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1)), 1):
                if cell.value == "Link":
                    for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        link_cell = row[0]
                        if link_cell.value and isinstance(link_cell.value, str) and link_cell.value.startswith("http"):
                            link_cell.hyperlink = link_cell.value
                            link_cell.style = "Hyperlink"
                    break

def highlight_new_rows(filtered_data, output_filename, timestamp, output_dir):
    """
    Highlight new rows in yellow in the 'Keyword Matches' Excel sheet based on PostedDate compared to the previous Excel file timestamp.

    Args:
        filtered_data (pd.DataFrame): DataFrame of filtered keyword matches (must include 'PostedDate').
        output_filename (str): Path to the output Excel file to highlight.
        timestamp (str): Current timestamp string used for output file naming.
        output_dir (str): Directory containing previous output Excel files.

    Returns:
        None

    Notes:
        - Compares PostedDate values to the timestamp of the previous Excel file in output_dir.
        - Highlights rows with PostedDate newer than the previous file's timestamp in yellow.
        - Logs the number of highlighted rows and any parse failures.
        - If no previous file is found, no rows are highlighted.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    import re as _re
    from dateutil.parser import parse as dtparse
    output_files = [f for f in os.listdir(output_dir) if f.endswith('_SAMOpportunities.xlsx')]
    output_files = sorted(output_files)
    if len(output_files) >= 2:
        prev_file = output_files[-2]
        match = _re.match(r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})_SAMOpportunities\.xlsx$", prev_file)
        if match:
            prev_timestamp_str = match.group(1)
            prev_timestamp_naive = datetime.strptime(prev_timestamp_str, "%Y-%m-%d_%H-%M-%S")
            prev_date = prev_timestamp_naive.date()
            new_row_indices = []
            parse_failures = 0
            failed_examples = []
            for idx, val in enumerate(filtered_data["PostedDate"].fillna("").astype(str)):
                date_str = val.strip()
                try:
                    row_dt = dtparse(date_str, fuzzy=True)
                except Exception:
                    try:
                        row_dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S.%f-%z")
                    except Exception:
                        parse_failures += 1
                        if len(failed_examples) < 5:
                            failed_examples.append(date_str)
                        continue
                if row_dt.date() >= prev_date:
                    new_row_indices.append(idx)
            wb = load_workbook(output_filename)
            ws = wb['Keyword Matches']
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for idx in new_row_indices:
                excel_row = idx + 2
                for cell in ws[excel_row]:
                    cell.fill = yellow_fill



            # Prepare DataFrame for 'New Opportunities' (drop specified columns)
            drop_cols = [col for col in ["Matched Keywords Count", "Description", "NoticeId"] if col in filtered_data.columns]
            new_opps_df = filtered_data.drop(columns=drop_cols)
            # Only keep new rows
            new_opps_df = new_opps_df.iloc[new_row_indices]

            # Create 'New Opportunities' sheet with only new rows and dropped columns
            if 'New Opportunities' in wb.sheetnames:
                del wb['New Opportunities']
            new_ws = wb.create_sheet('New Opportunities')

            # Write header
            for col_idx, col_name in enumerate(new_opps_df.columns, 1):
                new_ws.cell(row=1, column=col_idx, value=col_name)

            # Write new rows
            for row_num, row in enumerate(new_opps_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    new_ws.cell(row=row_num, column=col_idx, value=value)

            # Make 'Link' column a clickable hyperlink if present
            link_col_idx = None
            for col_idx, col_name in enumerate(new_opps_df.columns, 1):
                if col_name == "Link":
                    link_col_idx = col_idx
                    break
            if link_col_idx:
                for row in new_ws.iter_rows(min_row=2, min_col=link_col_idx, max_col=link_col_idx):
                    cell = row[0]
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"

            # Auto-size all columns to fit content
            for col_idx, col_name in enumerate(new_opps_df.columns, 1):
                max_length = len(str(col_name))
                for row in new_ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell_value = row[0].value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                new_ws.column_dimensions[new_ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

            wb.save(output_filename)
            logging.info(f"Highlighted {len(new_row_indices)} new rows in yellow based on PostedDate on or after previous file date: {prev_date}.")
            logging.info(f"Created 'New Opportunities' sheet with {len(new_row_indices)} rows.")
            if parse_failures > 0:
                logging.info(f"Could not parse PostedDate for {parse_failures} rows. Examples: {failed_examples}")
        else:
            logging.info(f"Could not extract timestamp from previous Excel filename for highlighting: {prev_file}")
    else:
        logging.info("No previous Excel file found for row comparison/highlighting.")

def process_csv_file(csv_filename, timestamp):
    """
    Process the downloaded CSV file by filtering, cleaning, and exporting results to Excel with highlighting.

    Args:
        csv_filename (str): Path to the input CSV file to process.
        timestamp (str): Timestamp string used for output file naming and logging.

    Returns:
        None

    Steps:
        1. Reads the CSV file into a DataFrame.
        2. Filters rows by organizations from cfg/keyorgs.txt.
        3. Filters rows by keywords from cfg/keywords.txt, adds match columns.
        4. Prepares keyword match counts for a summary sheet.
        5. Filters columns based on cfg/keep_cols.txt.
        6. Cleans DataFrame column names and string values.
        7. Writes results to an Excel file with formatting and hyperlinks.
        8. Highlights new rows based on PostedDate compared to previous output.
    """
    try:
        logging.info("\nReading CSV file...")
        raw_data = pd.read_csv(csv_filename, encoding='ISO-8859-1', dtype=str, low_memory=False)
        logging.info(f"CSV file loaded successfully. Rows: {len(raw_data)}, Columns: {len(raw_data.columns)}")

        # Organization filtering
        keyorgs_file = os.path.join(os.path.dirname(os.path.dirname(csv_filename)), "cfg", "keyorgs.txt")
        try:
            with open(keyorgs_file, "r", encoding="utf-8") as kof:
                keyorgs = [line.strip() for line in kof if line.strip()]
            logging.info(f"Loaded {len(keyorgs)} organization filters from keyorgs.txt.")
        except Exception as e:
            logging.info(f"Failed to read keyorgs from {keyorgs_file}: {e}")
            keyorgs = []
        org_filtered_data = filter_by_organization(raw_data, keyorgs)
        logging.info(f"{len(org_filtered_data)} rows after organization filtering.")

        # Keyword filtering
        keywords_file = os.path.join(os.path.dirname(os.path.dirname(csv_filename)), "cfg", "keywords.txt")
        try:
            with open(keywords_file, "r", encoding="utf-8") as kf:
                keywords = [line.strip() for line in kf if line.strip()]
            logging.info(f"\nLoaded {len(keywords)} keywords for filtering from keywords.txt.")
        except Exception as e:
            logging.info(f"Failed to read keywords from {keywords_file}: {e}")
            keywords = []
        filtered_data = filter_by_keywords(org_filtered_data, keywords)
        logging.info(f"{len(filtered_data)} rows after keyword filtering.")

        # Prepare keyword match counts for the 'Keyword Match Counts' sheet
        keyword_counts = []
        if keywords:
            for k in keywords:
                pattern, flags = make_pattern_and_flags(k)
                mask = org_filtered_data['Description'].str.contains(pattern, case=False, na=False, regex=True, flags=flags)
                count = mask.sum()
                keyword_counts.append({'Keyword': k, 'Count': int(count)})
            keyword_counts_df = pd.DataFrame(keyword_counts)
            keyword_counts_df = keyword_counts_df.sort_values(by='Count', ascending=False).reset_index(drop=True)
        else:
            keyword_counts_df = pd.DataFrame(columns=['Keyword', 'Count'])

        # Column filtering
        keep_cols_file = os.path.join(os.path.dirname(os.path.dirname(csv_filename)), "cfg", "keep_cols.txt")
        try:
            with open(keep_cols_file, "r", encoding="utf-8") as kcf:
                keep_cols = [line.strip() for line in kcf if line.strip()]
            filtered_data = filter_columns(filtered_data, keep_cols)
            logging.info(f"Keeping {len(keep_cols)} columns as specified in keep_cols.txt (plus Matched Keywords columns).")
        except Exception as e:
            logging.info(f"Failed to read or apply keep_cols.txt: {e}")

        # Data cleaning
        filtered_data = clean_dataframe(filtered_data)

        # Save to Excel
        logging.info("Saving Excel file...")
        output_dir = os.path.join(os.path.dirname(csv_filename), '..', 'output')
        output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)
        output_filename = os.path.join(output_dir, f"{timestamp}_SAMOpportunities.xlsx")
        write_to_excel(filtered_data, keyword_counts_df, output_filename)

        # Highlight new rows
        highlight_new_rows(filtered_data, output_filename, timestamp, output_dir)

        logging.info(f"Processed file saved as: {output_filename}")
    except Exception as e:
        logging.info(f"Failed to read CSV file: {e}")

def main():
    """
    Main entry point for the SAM.gov webscrape process.

    Sets up logging, prepares input/output directories, reads the download URL, downloads the CSV,
    and processes the CSV file through filtering, cleaning, and Excel export.

    Returns:
        None

    Steps:
        1. Sets up logging to both file and console.
        2. Determines script directory (handles frozen executable or script).
        3. Prepares logs, input, and output directories.
        4. Reads the download URL from configuration.
        5. Downloads the CSV file to the input directory.
        6. Processes the CSV file (filter, clean, export, highlight).
    """
    # Set up logging to both file and console
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    logs_dir = os.path.join(script_dir, "logs")
    os.makedirs(logs_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    log_filename = os.path.join(logs_dir, f"{timestamp}_webscrape.log")

    logging.basicConfig(
        level=logging.INFO,
        format="%(message)s",
        handlers=[
            logging.FileHandler(log_filename, encoding="utf-8"),
            logging.StreamHandler(sys.stdout)
        ]
    )

    logging.info("SAM.gov Webscrape")
    logging.info("=================")
    input_folder, output_folder = setup_directories(script_dir)
    url = read_url_from_file(script_dir)
    if not url:
        return
    csv_filename = os.path.join(input_folder, f"ContractOpportunitiesFullCSV.csv")
    download_csv_file(url, csv_filename)
    process_csv_file(csv_filename, timestamp)

if __name__ == "__main__":
    main()